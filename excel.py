from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.workbook.views import BookView

def write_to_excel(input_txt, output_xlsx):
    # with open(input_txt + '.txt', 'r') as input_file:
    #     with open('temp.txt', 'a+') as temp_file:
    #         row = 1
    #         for line in input_file:
    #             final_string = ''
    #             for field in line.split('"'):
    #                 field = field.strip()
    #                 field += '"'
    #                 final_string += field
    #
    #             temp_file.write(final_string + "\n")
    #             row += 1


    with open(input_txt + '.txt', 'r') as input_file:
        wb = Workbook()
        ws = wb.active
        row = 1

        for line in input_file:
            for index, field in enumerate(line.split('"')):
                print(row)
                ws[get_column_letter(index+1) + str(row)] = field
            row += 1
        wb.save(output_xlsx + '.xlsx')

write_to_excel('full_data', 'full_data')